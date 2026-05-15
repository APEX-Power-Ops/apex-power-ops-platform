import base64
import json

from openpyxl import Workbook

from app.project_seed_sources import (
    DEFAULT_ESTIMATOR_WORKBOOK_NAME,
    clear_project_seed_cache,
    extract_topology_labels_from_text,
    load_project_seed_sources,
)


def _make_token() -> dict[str, str]:
    payload = {
        "actor_id": "tech-001",
        "actor_role": "field_tech",
        "project_scope": ["proj-001"],
    }
    encoded = base64.b64encode(json.dumps(payload).encode()).decode()
    return {"Authorization": f"Bearer {encoded}"}


def _write_estimator_workbook(path):
    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = "Updated"
        sheet.append([])
        sheet.append([None, "Updated"])
        sheet.append([None, None, "NETA", None, "Miner Temp Power", "SLD: E01-00, E01-01, E01-02"])
        sheet.append([None, None, "ATS", None, "Santa Teresa, NM", "Dated: 03/05/2026"])
        sheet.append([None, None, "QTY", "Section", "Apparatus Type", "Designation", "Notes", None, "Hrs/Unit", "Hrs/Line"])
        sheet.append([None, None, 3, 7.5, "Switch MV - Fused Disconnect", "Pole Disconnect", "E01-00", None, 2.5, 7.5])
        sheet.append([None, None, 2, 7.1, "Switchboard - Distribution LV", "SWBD", "E01-01", None, 3, 6])
        workbook.save(path)
    finally:
        workbook.close()


def _write_scope_estimator_workbook(path):
    workbook = Workbook()
    try:
        reference_sheet = workbook.active
        reference_sheet.title = "Equipment Reference"
        reference_sheet["L4"] = "A1) Medium-Voltage - Core"
        reference_sheet["M4"] = 91190.625
        reference_sheet["L5"] = "10.X"
        reference_sheet["M5"] = 0

        scope_sheet = workbook.create_sheet("A1) Medium-Voltage - Core")
        scope_sheet["C4"] = "ATS"
        scope_sheet["J3"] = 10.5
        scope_sheet["M4"] = 1
        scope_sheet["P3"] = 91190.625
        scope_sheet["E6"] = "Medium Voltage"
        scope_sheet["C7"] = 2
        scope_sheet["E7"] = "Transformer MV"
        scope_sheet["I7"] = 3.5
        scope_sheet["J7"] = 7
        scope_sheet["C8"] = 1
        scope_sheet["E8"] = "Switchgear MV"
        scope_sheet["I8"] = 2
        scope_sheet["J8"] = 2
        workbook.save(path)
    finally:
        workbook.close()


def test_extract_topology_labels_from_text_finds_equipment_labels():
    text = "MVTX -A SWBD -C PWR SKID -25 LVPP -2 MVS -4"

    labels = extract_topology_labels_from_text(text)

    assert labels == ["PWR SKID -25", "MVTX -A", "SWBD -C", "MVS -4", "LVPP -2"]


def test_load_project_seed_sources_reads_workbook_line_items(tmp_path):
    workbook_path = tmp_path / "estimator.xlsm"
    _write_estimator_workbook(workbook_path)

    data = load_project_seed_sources(str(workbook_path), str(tmp_path / "missing.pdf"))

    assert data["project_name"] == "Miner Temp Power"
    assert data["location"] == "Santa Teresa, NM"
    assert data["source_sheet"] == "Updated"
    assert len(data["line_items"]) == 2
    assert len(data["expanded_apparatus_candidates"]) == 5
    assert data["expanded_apparatus_candidates"][0]["display_name"] == "Pole Disconnect 01"
    assert data["expanded_apparatus_candidates"][3]["display_name"] == "SWBD 01"


def test_load_project_seed_sources_reads_scope_sheet_estimator(tmp_path):
    workbook_path = tmp_path / "building-ab-estimator.xlsm"
    _write_scope_estimator_workbook(workbook_path)

    data = load_project_seed_sources(str(workbook_path), str(tmp_path / "missing.pdf"))

    assert data["project_name"] == "building-ab-estimator"
    assert data["source_sheet"] == "scope_sheets"
    assert data["source_format"] == "scope_sheets"
    assert data["scope_count"] == 1
    assert data["scope_sheets"] == ["A1) Medium-Voltage - Core"]
    assert data["metadata"]["estimator_source_format"] == "scope_sheets"
    assert data["metadata"]["estimator_scope_count"] == 1
    assert len(data["line_items"]) == 2
    assert data["line_items"][0]["section"] == "Medium Voltage"
    assert data["line_items"][0]["scope_sheet"] == "A1) Medium-Voltage - Core"
    assert data["line_items"][0]["scope_type"] == "ATS"
    assert data["line_items"][0]["scope_quoted_amount"] == 91190.625
    assert len(data["expanded_apparatus_candidates"]) == 3
    assert data["expanded_apparatus_candidates"][0]["display_name"] == "Transformer MV 01"
    assert data["expanded_apparatus_candidates"][2]["display_name"] == "Switchgear MV"


def test_load_project_seed_sources_uses_project_miner_planning_root(monkeypatch, tmp_path):
    planning_root = tmp_path / "Project Miner PM Planning"
    planning_root.mkdir()
    workbook_path = planning_root / DEFAULT_ESTIMATOR_WORKBOOK_NAME
    _write_estimator_workbook(workbook_path)

    monkeypatch.delenv("APEX_PROJECT_ESTIMATOR_WORKBOOK", raising=False)
    monkeypatch.delenv("APEX_PROJECT_SLD_PDF", raising=False)
    monkeypatch.setenv("APEX_PROJECT_MINER_PLANNING_ROOT", str(planning_root))
    clear_project_seed_cache()

    data = load_project_seed_sources()

    assert data["metadata"]["estimator_workbook_path"] == str(workbook_path)
    assert data["metadata"]["estimator_workbook_found"] is True
    assert data["project_name"] == "Miner Temp Power"
    assert len(data["expanded_apparatus_candidates"]) == 5


def test_project_apparatus_plan_route_returns_workbook_backed_rows(client, monkeypatch, tmp_path):
    workbook_path = tmp_path / "estimator.xlsm"
    _write_estimator_workbook(workbook_path)

    monkeypatch.setenv("APEX_PROJECT_ESTIMATOR_WORKBOOK", str(workbook_path))
    monkeypatch.setenv("APEX_PROJECT_SLD_PDF", str(tmp_path / "missing.pdf"))
    clear_project_seed_cache()

    response = client.get("/api/v1/reads/project-apparatus-plan", headers=_make_token())

    assert response.status_code == 200
    data = response.json()
    assert data["project_name"] == "Miner Temp Power"
    assert data["metadata"]["estimator_workbook_found"] is True
    assert data["metadata"]["sld_pdf_found"] is False
    assert len(data["line_items"]) == 2
    assert len(data["expanded_apparatus_candidates"]) == 5
