import base64
import json

from openpyxl import Workbook

from app.project_import_candidate import clear_project_import_candidate_cache, load_project_import_candidate
from app.project_seed_sources import clear_project_seed_cache
from app.project_tracker_sources import clear_project_tracker_cache
from app.seed_workbooks import clear_seed_cache


def _make_token() -> dict[str, str]:
    payload = {
        "actor_id": "pm-001",
        "actor_role": "pm",
        "project_scope": ["proj-001"],
    }
    encoded = base64.b64encode(json.dumps(payload).encode()).decode()
    return {"Authorization": f"Bearer {encoded}"}


def _clear_all_caches() -> None:
    clear_project_import_candidate_cache()
    clear_project_seed_cache()
    clear_project_tracker_cache()
    clear_seed_cache()


def _write_candidate_estimator(path):
    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = "Updated"
        sheet.append([])
        sheet.append([None, "Updated"])
        sheet.append([None, None, "NETA", None, "Miner Temp Power", "SLD: E01-00, E01-01"])
        sheet.append([None, None, "ATS", None, "Santa Teresa, NM", "Dated: 03/05/2026"])
        sheet.append([None, None, "QTY", "Section", "Apparatus Type", "Designation", "Notes", None, "Hrs/Unit", "Hrs/Line"])
        sheet.append([None, None, 2, "Temp Power", "Switch MV - Fused Disconnect", "PD-1", "E01-00", None, 2.5, 5])
        sheet.append([None, None, 1, "Temp Power", "Switch MV - Fused Disconnect", "PD-1", "E01-00", None, 2.5, 2.5])
        sheet.append([None, None, 1, "Temp Power", "Panelboard LV", None, None, None, 1.5, 1.5])
        workbook.save(path)
    finally:
        workbook.close()


def _write_ground_resistance_lot_estimator(path):
    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = "Updated"
        sheet.append([])
        sheet.append([None, "Updated"])
        sheet.append([None, None, "NETA", None, "Miner Temp Power", "SLD: E01-00, E01-01"])
        sheet.append([None, None, "ATS", None, "Santa Teresa, NM", "Dated: 03/05/2026"])
        sheet.append([None, None, "QTY", "Section", "Apparatus Type", "Designation", "Notes", None, "Hrs/Unit", "Hrs/Line"])
        sheet.append(
            [
                None,
                None,
                3,
                "7.13",
                "Ground Resistance Test - Two-Point (Lot)",
                None,
                "E01-00, E01-01, E01-02",
                None,
                8,
                24,
            ]
        )
        workbook.save(path)
    finally:
        workbook.close()


def _write_tracker_formula_cache_break_workbook(path):
    workbook = Workbook()
    try:
        project_form = workbook.active
        project_form.title = "Project_Form"
        project_form["B4"] = "Client:"
        project_form["C4"] = "RESA"
        project_form["B5"] = "Project:"
        project_form["C5"] = "Tracker Project"
        project_form["E4"] = "Scope_5"

        task_entry = workbook.create_sheet("Task_Entry")
        task_entry.append(
            [
                "Scope",
                "NETA_Standard",
                "Task_ID",
                "Task",
                "Apparatus",
                "Designation",
                "Drawing",
                "Apparatus_Hourrs",
            ]
        )
        task_entry.append(["Scope_5", "ATS", "5.1.1", "SWGR-GS", "Switchboard - Low Voltage", "SWGR-GS", None, 4])
        task_entry.append(["Scope_5", "ATS", "5.1.2", None, "Circuit Breaker LV - EO (LSIG)", "GEN-A", None, 5])

        all_tasks = workbook.create_sheet("All_Tasks")
        all_tasks.append(
            [
                "Scope",
                "NETA_Standard",
                "Task_ID",
                "Task",
                "Apparatus",
                "Designation",
                "Drawing",
                "Date Due",
                "Notes",
                "Assessment",
                "DATASHEET",
                "DATE COMPLETED",
                "NOTES2",
                "% COMPLETION",
                "TASK DELAYS",
                "Apparatus Hours",
                "Remaining Hours",
                "ACTUAL HOURS",
                "STATUS",
                "AVAILABILITY",
                "PRIORITY",
                "Apparatus Category",
            ]
        )
        for task_id, apparatus, designation in [
            ("5.1.1", "Switchboard - Low Voltage", "SWGR-GS"),
            ("5.1.2", "Circuit Breaker LV - EO (LSIG)", "GEN-A"),
        ]:
            all_tasks.append(
                [
                    "Scope_5",
                    "ATS",
                    task_id,
                    "SWGR-GS",
                    apparatus,
                    designation,
                    "#REF!",
                    "#REF!",
                    "#REF!",
                    "#REF!",
                    "#REF!",
                    "#REF!",
                    "#REF!",
                    "#REF!",
                    "#REF!",
                    "#REF!",
                    "#REF!",
                    "#REF!",
                    "#REF!",
                    "#REF!",
                    "#REF!",
                    None,
                ]
            )
        workbook.save(path)
    finally:
        workbook.close()


def _write_clean_reference_tracker_workbook(path):
    workbook = Workbook()
    try:
        project_form = workbook.active
        project_form.title = "Project_Form"
        project_form["B4"] = "Client:"
        project_form["C4"] = "Garney"
        project_form["B5"] = "Project:"
        project_form["C5"] = "Central Mesa"
        project_form["E4"] = "Scope_5"

        task_entry = workbook.create_sheet("Task_Entry")
        task_entry.append(
            [
                "Scope",
                "NETA_Standard",
                "Task_ID",
                "Task",
                "Apparatus",
                "Designation",
                "Drawing",
                "Apparatus_Hourrs",
            ]
        )
        task_entry.append(["Scope_5", "ATS", "5.1.1", "SWGR-GS", "Switchboard - Low Voltage", "SWGR-GS", "E01", 4])
        task_entry.append(["Scope_5", "ATS", "5.1.2", "GEN-A", "Circuit Breaker LV - EO (LSIG)", "GEN-A", "E01", 5])

        all_tasks = workbook.create_sheet("All_Tasks")
        all_tasks.append(
            [
                "Scope",
                "NETA_Standard",
                "Task_ID",
                "Task",
                "Apparatus",
                "Designation",
                "Drawing",
                "Date Due",
                "Notes",
                "Assessment",
                "DATASHEET",
                "DATE COMPLETED",
                "NOTES2",
                "% COMPLETION",
                "TASK DELAYS",
                "Apparatus Hours",
                "Remaining Hours",
                "ACTUAL HOURS",
                "STATUS",
                "AVAILABILITY",
                "PRIORITY",
                "Apparatus Category",
            ]
        )
        all_tasks.append(
            [
                "Scope_5",
                "ATS",
                "5.1.1",
                "SWGR-GS",
                "Switchboard - Low Voltage",
                "SWGR-GS",
                "E01",
                None,
                None,
                None,
                False,
                None,
                None,
                100,
                None,
                4,
                0,
                4,
                "COMPLETED",
                "READY",
                "HIGH",
                "Switchboard",
            ]
        )
        all_tasks.append(
            [
                "Scope_5",
                "ATS",
                "5.1.2",
                "GEN-A",
                "Circuit Breaker LV - EO (LSIG)",
                "GEN-A",
                "E01",
                None,
                None,
                None,
                False,
                None,
                None,
                0,
                None,
                5,
                5,
                0,
                "NOT STARTED",
                "READY",
                "MEDIUM",
                "CB Primary",
            ]
        )
        workbook.save(path)
    finally:
        workbook.close()


def test_load_project_import_candidate_groups_tasks_and_preserves_traceability(tmp_path):
    workbook_path = tmp_path / "estimator.xlsm"
    _write_candidate_estimator(workbook_path)
    _clear_all_caches()

    candidate = load_project_import_candidate(str(workbook_path), str(tmp_path / "missing.pdf"))

    assert candidate["candidate_id"] == "pm-import-candidate-miner-temp-power"
    assert candidate["mutation_authority"] == "not_admitted"
    assert candidate["review_status"] == "draft_review_only"
    assert candidate["summary"]["workpackage_count"] == 1
    assert candidate["summary"]["task_count"] == 3
    assert candidate["summary"]["apparatus_candidate_count"] == 4
    assert candidate["project"]["location"] == "Santa Teresa, NM"
    assert candidate["workpackages"][0]["task_count"] == 3
    assert candidate["workpackages"][0]["apparatus_candidate_count"] == 4
    assert candidate["workpackages"][0]["planned_hours"] == 9.0
    assert candidate["workpackages"][0]["tasks"][0]["source_ref"]["source_row"] == 6
    assert candidate["workpackages"][0]["tasks"][0]["apparatus_candidates"][0]["source_row"] == 6
    warning_codes = {warning["code"] for warning in candidate["warnings"]}
    assert "MISSING_SLD_PDF" in warning_codes
    assert "DUPLICATE_LINE_ITEM_GROUPS" in warning_codes
    assert "MISSING_DESIGNATIONS" in warning_codes
    assert "MISSING_DRAWING_REFS" in warning_codes
    assert candidate["summary"]["human_decision_count"] >= 2
    assert "write_supabase" in candidate["review_guidance"]["not_allowed_now"]
    freshness = candidate["source_freshness"]
    assert freshness["strategy"] == "path_size_mtime_fingerprint"
    assert freshness["mutation_authority"] == "not_admitted"
    assert freshness["missing_count"] >= 1
    source_files = {source["source_id"]: source for source in freshness["source_files"]}
    assert source_files["estimator_workbook"]["found"] is True
    assert source_files["estimator_workbook"]["fingerprint"]
    assert source_files["estimator_workbook"]["modified_at"].endswith("Z")
    assert source_files["sld_pdf"]["freshness_status"] == "missing"


def test_ground_resistance_lot_gets_pm_designation_without_expanding_measurements(tmp_path):
    workbook_path = tmp_path / "estimator.xlsm"
    _write_ground_resistance_lot_estimator(workbook_path)
    _clear_all_caches()

    candidate = load_project_import_candidate(str(workbook_path), str(tmp_path / "missing.pdf"))

    task = candidate["workpackages"][0]["tasks"][0]
    assert task["source_line_id"] == "miner-line-001"
    assert task["designation"] == "Ground Resistance Test Lot"
    assert task["quantity"] == 3
    assert task["planned_hours"] == 24.0
    assert len(task["apparatus_candidates"]) == 1
    assert task["apparatus_candidates"][0]["designation"] == "Ground Resistance Test Lot"
    assert task["apparatus_candidates"][0]["display_name"] == "Ground Resistance Test Lot"
    assert task["apparatus_candidates"][0]["planned_hours"] == 24
    assert candidate["summary"]["apparatus_candidate_count"] == 1
    warning_codes = {warning["code"] for warning in candidate["warnings"]}
    assert "MISSING_DESIGNATIONS" not in warning_codes


def test_project_import_candidate_route_returns_read_only_candidate(client, monkeypatch, tmp_path):
    workbook_path = tmp_path / "estimator.xlsm"
    _write_candidate_estimator(workbook_path)
    monkeypatch.setenv("APEX_PROJECT_ESTIMATOR_WORKBOOK", str(workbook_path))
    monkeypatch.setenv("APEX_PROJECT_SLD_PDF", str(tmp_path / "missing.pdf"))
    _clear_all_caches()

    response = client.get("/api/v1/reads/project-import-candidate", headers=_make_token())

    assert response.status_code == 200
    candidate = response.json()
    assert candidate["mutation_authority"] == "not_admitted"
    assert candidate["summary"]["task_count"] == 3
    assert candidate["summary"]["apparatus_candidate_count"] == 4
    assert candidate["workpackages"][0]["tasks"][0]["source_ref"]["estimator_workbook_path"] == str(workbook_path)
    assert candidate["source_freshness"]["source_files"][0]["source_id"] == "estimator_workbook"


def test_project_import_candidate_flags_all_tasks_formula_cache_break(tmp_path, monkeypatch):
    workbook_path = tmp_path / "estimator.xlsm"
    data_entry_workbook = tmp_path / "data-entry.xlsm"
    reference_tracker_workbook = tmp_path / "reference-tracker.xlsm"
    _write_candidate_estimator(workbook_path)
    _write_tracker_formula_cache_break_workbook(data_entry_workbook)
    _write_clean_reference_tracker_workbook(reference_tracker_workbook)

    monkeypatch.setenv("APEX_PROJECT_DATA_ENTRY_WORKBOOK", str(data_entry_workbook))
    monkeypatch.setenv("APEX_REFERENCE_TRACKER_WORKBOOK", str(reference_tracker_workbook))
    _clear_all_caches()

    candidate = load_project_import_candidate(str(workbook_path), str(tmp_path / "missing.pdf"))

    warning = next(warning for warning in candidate["warnings"] if warning["code"] == "PROJECT_DATA_ENTRY_FORMULA_ERRORS")
    assert warning["formula_error_pattern"] == "all_tasks_formula_cache_break"
    assert "All 2 All_Tasks row(s)" in warning["message"]
    assert "Task_Entry source rows are still present" in warning["message"]
    assert "BuildAll/PopulateAllTasks" in warning["review_action"]
    assert warning["formula_error_vba_lineage_modules"] == ["BuildAll", "PopulateAllTasks_FromSheets"]
    assert "Reference workbook example: reference-tracker.xlsm currently loads without formula errors" in warning[
        "formula_error_pattern_detail"
    ]
    assert "2 All_Tasks row(s)" in warning["formula_error_pattern_detail"]
    assert "2 Task_Entry row(s)" in warning["formula_error_pattern_detail"]
    warning_codes = {item["code"] for item in candidate["warnings"]}
    assert "REFERENCE_TRACKER_FORMULA_ERRORS" not in warning_codes
