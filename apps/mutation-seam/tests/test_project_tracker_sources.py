from openpyxl import Workbook

from app.project_tracker_sources import (
    DEFAULT_PROJECT_DATA_ENTRY_WORKBOOK_NAME,
    DEFAULT_REFERENCE_TRACKER_WORKBOOK_NAME,
    clear_project_tracker_cache,
    load_project_tracker_sources,
)


def _write_tracker_workbook(path, *, client="RESA", project="Tracker Project", status="NOT STARTED"):
    workbook = Workbook()
    try:
        project_form = workbook.active
        project_form.title = "Project_Form"
        project_form["B4"] = "Client:"
        project_form["C4"] = client
        project_form["B5"] = "Project:"
        project_form["C5"] = project
        project_form["B6"] = "Job #:"
        project_form["C6"] = 123456
        project_form["E4"] = "Scope_1"

        task_entry = workbook.create_sheet("Task_Entry")
        task_entry.append(
            [
                "Scope",
                "NETA_Standard",
                "Task_ID",
                "Task",
                "Apparatus",
                "Designation",
                "Drawing",
                "Apparatus_Hourrs",
            ]
        )
        task_entry.append(["Scope_1", "ATS", "1.1.1", "SWGR-A", "Switchboard - Low Voltage", "SWGR-A", "E01", 4])
        task_entry.append(["Scope_1", "ATS", "1.1.2", None, "Circuit Breaker LV - EO (LSIG)", "CB-1", "E01", 5])

        all_tasks = workbook.create_sheet("All_Tasks")
        all_tasks.append(
            [
                "Scope",
                "NETA_Standard",
                "Task_ID",
                "Task",
                "Apparatus",
                "Designation",
                "Drawing",
                "Date Due",
                "Notes",
                "Assessment",
                "DATASHEET",
                "DATE COMPLETED",
                "NOTES2",
                "% COMPLETION",
                "TASK DELAYS",
                "Apparatus Hours",
                "Remaining Hours",
                "ACTUAL HOURS",
                "STATUS",
                "AVAILABILITY",
                "PRIORITY",
                "Apparatus Category",
            ]
        )
        all_tasks.append(
            [
                "Scope_1",
                "ATS",
                "1.1.1",
                "SWGR-A",
                "Switchboard - Low Voltage",
                "SWGR-A",
                "E01",
                None,
                None,
                None,
                False,
                None,
                None,
                0,
                None,
                4,
                4,
                0,
                status,
                "READY",
                None,
                "Switchboard",
            ]
        )
        all_tasks.append(
            [
                "Scope_1",
                "ATS",
                "1.1.2",
                "SWGR-A",
                "Circuit Breaker LV - EO (LSIG)",
                "CB-1",
                "#REF!",
                "#REF!",
                None,
                None,
                False,
                None,
                None,
                0,
                None,
                "#REF!",
                "#REF!",
                0,
                "#REF!",
                "#REF!",
                None,
                "CB Primary",
            ]
        )
        workbook.save(path)
    finally:
        workbook.close()


def test_load_project_tracker_sources_reads_task_entry_and_all_tasks(tmp_path):
    data_entry = tmp_path / "data-entry.xlsm"
    reference_tracker = tmp_path / "reference-tracker.xlsm"
    _write_tracker_workbook(data_entry, client="Blank", project="Template", status="NOT STARTED")
    _write_tracker_workbook(reference_tracker, client="Garney", project="Central Mesa", status="COMPLETED")

    data = load_project_tracker_sources(str(data_entry), str(reference_tracker))

    assert data["metadata"]["project_data_entry_workbook_found"] is True
    assert data["metadata"]["reference_tracker_workbook_found"] is True
    assert data["project_data_entry"]["project_form"]["Client"] == "Blank"
    assert data["project_data_entry"]["workscope_names"] == ["Scope_1"]
    assert data["project_data_entry"]["task_entry_count"] == 2
    assert data["project_data_entry"]["all_tasks_count"] == 2
    assert data["project_data_entry"]["task_entry_scope_counts"] == {"Scope_1": 2}
    assert data["project_data_entry"]["status_counts"] == {"NOT STARTED": 1, "#REF!": 1}
    assert data["project_data_entry"]["formula_error_row_count"] == 1
    assert data["project_data_entry"]["formula_error_cell_count"] == 6
    assert data["project_data_entry"]["formula_error_column_counts"] == {
        "Drawing": 1,
        "Date Due": 1,
        "Apparatus Hours": 1,
        "Remaining Hours": 1,
        "STATUS": 1,
        "AVAILABILITY": 1,
    }
    assert data["project_data_entry"]["formula_error_sample_rows"] == [
        {
            "source_row": 3,
            "error_columns": [
                "Drawing",
                "Date Due",
                "Apparatus Hours",
                "Remaining Hours",
                "STATUS",
                "AVAILABILITY",
            ],
            "scope": "Scope_1",
            "task_id": "1.1.2",
            "task": "SWGR-A",
            "apparatus": "Circuit Breaker LV - EO (LSIG)",
            "designation": "CB-1",
        }
    ]
    assert data["project_data_entry"]["task_entry_sample"][0]["apparatus_hours"] == 4
    assert data["reference_tracker"]["project_form"]["Project"] == "Central Mesa"
    assert data["reference_tracker"]["status_counts"] == {"COMPLETED": 1, "#REF!": 1}


def test_load_project_tracker_sources_uses_project_miner_planning_root(monkeypatch, tmp_path):
    planning_root = tmp_path / "Project Miner PM Planning"
    planning_root.mkdir()
    data_entry = planning_root / DEFAULT_PROJECT_DATA_ENTRY_WORKBOOK_NAME
    reference_tracker = planning_root / DEFAULT_REFERENCE_TRACKER_WORKBOOK_NAME
    _write_tracker_workbook(data_entry)
    _write_tracker_workbook(reference_tracker, client="Garney")

    monkeypatch.setenv("APEX_PROJECT_MINER_PLANNING_ROOT", str(planning_root))
    monkeypatch.delenv("APEX_PROJECT_DATA_ENTRY_WORKBOOK", raising=False)
    monkeypatch.delenv("APEX_REFERENCE_TRACKER_WORKBOOK", raising=False)
    clear_project_tracker_cache()

    data = load_project_tracker_sources()

    assert data["metadata"]["project_data_entry_workbook_path"] == str(data_entry)
    assert data["metadata"]["reference_tracker_workbook_path"] == str(reference_tracker)
    assert data["project_data_entry"]["found"] is True
    assert data["reference_tracker"]["project_form"]["Client"] == "Garney"
