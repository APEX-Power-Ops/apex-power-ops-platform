import base64
import json

from openpyxl import Workbook

from app.seed_workbooks import (
    DEFAULT_CAPABILITY_WORKBOOK_NAME,
    DEFAULT_EQUIPMENT_WORKBOOK_NAME,
    clear_seed_cache,
    load_seed_data,
)


def _make_token(actor_id: str = "tech-001", actor_role: str = "field_tech") -> dict[str, str]:
    payload = {
        "actor_id": actor_id,
        "actor_role": actor_role,
        "project_scope": ["proj-001"],
    }
    encoded = base64.b64encode(json.dumps(payload).encode()).decode()
    return {"Authorization": f"Bearer {encoded}"}


def _write_equipment_workbook(path):
    workbook = Workbook()
    try:
        all_equipment = workbook.active
        all_equipment.title = "ALL Equipment"
        all_equipment.append(["CATEGORY", "ID# BOB", "EQUIPMENT", "MFR.", "MODEL", "SERIAL #", "LAST CAL DATE", "CAL DATE '26", "ASSIGNED", "NOTES"])
        all_equipment.append(["TEST EQUIP", "PHX-MEG5-01", "Megohmeter - 5kVDC", "Megger", "MIT525", "SN-001", "2025-11-01", "2026-11-01", "Ace Randolph", "Primary kit"])

        standard_tech_list = workbook.create_sheet("Standard Tech List")
        standard_tech_list.append(["CATEGORY", "PHX-ID-#", "EQUIPMENT", "Standard/Optional", "Ace Randolph", "Aaron Carter", "Name"])
        standard_tech_list.append(["TEST EQUIP", "PHX-MEG5-", "Megohmeter - 5kVDC", "Standard", "Yes", "Yes", None])
        workbook.save(path)
    finally:
        workbook.close()


def _write_capability_workbook(path):
    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = "Tech Capability"
        sheet.append([None, None, "Total Experience Score:", 101, 71])
        sheet.append([None, None, None, 101, 71])
        sheet.append(["TYPE", "Voltage", "TEST", "Ace Randolph (2)", "Aaron Carter (2)"])
        sheet.append(["IR", "LV", "Infrared Scan (Energized)", 3, 2, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 3, "Can Teach Others"])
        sheet.append(["ATS", "LV", "ATS Controller Swap", 2, 0, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 0, "No Experience"])
        workbook.save(path)
    finally:
        workbook.close()


def test_reads_crew_uses_workbook_backed_seed(client, monkeypatch, tmp_path):
    equipment_path = tmp_path / "equipment.xlsx"
    capability_path = tmp_path / "capability.xlsx"
    _write_equipment_workbook(equipment_path)
    _write_capability_workbook(capability_path)

    monkeypatch.setenv("APEX_FIELD_SEED_EQUIPMENT_WORKBOOK", str(equipment_path))
    monkeypatch.setenv("APEX_FIELD_SEED_CAPABILITY_WORKBOOK", str(capability_path))
    clear_seed_cache()

    response = client.get("/api/v1/reads/crew", headers=_make_token())

    assert response.status_code == 200
    data = response.json()
    assert data[0]["id"] == "tech-001"
    assert data[0]["name"] == "Ace Randolph"
    assert data[0]["experience_score"] == 101
    assert data[1]["name"] == "Aaron Carter"


def test_reads_inventory_and_capabilities_use_workbook_seed(client, monkeypatch, tmp_path):
    equipment_path = tmp_path / "equipment.xlsx"
    capability_path = tmp_path / "capability.xlsx"
    _write_equipment_workbook(equipment_path)
    _write_capability_workbook(capability_path)

    monkeypatch.setenv("APEX_FIELD_SEED_EQUIPMENT_WORKBOOK", str(equipment_path))
    monkeypatch.setenv("APEX_FIELD_SEED_CAPABILITY_WORKBOOK", str(capability_path))
    clear_seed_cache()

    inventory_response = client.get("/api/v1/reads/equipment-inventory", headers=_make_token())
    capability_response = client.get("/api/v1/reads/tech-capabilities", headers=_make_token())

    assert inventory_response.status_code == 200
    inventory = inventory_response.json()
    assert inventory[0]["equipment"] == "Megohmeter - 5kVDC"
    assert inventory[0]["assigned_to_name"] == "Ace Randolph"
    assert inventory[0]["assigned_to_id"] == "tech-001"

    assert capability_response.status_code == 200
    capabilities = capability_response.json()
    assert capabilities["score_scale"][0] == {"score": 0, "label": "No Experience"}
    assert capabilities["capabilities"][0]["test"] == "Infrared Scan (Energized)"
    assert capabilities["capabilities"][0]["scores"]["tech-001"] == 3


def test_workbook_seed_uses_project_miner_planning_root(monkeypatch, tmp_path):
    planning_root = tmp_path / "Project Miner PM Planning"
    planning_root.mkdir()
    equipment_path = planning_root / DEFAULT_EQUIPMENT_WORKBOOK_NAME
    capability_path = planning_root / DEFAULT_CAPABILITY_WORKBOOK_NAME
    _write_equipment_workbook(equipment_path)
    _write_capability_workbook(capability_path)

    monkeypatch.delenv("APEX_FIELD_SEED_EQUIPMENT_WORKBOOK", raising=False)
    monkeypatch.delenv("APEX_FIELD_SEED_CAPABILITY_WORKBOOK", raising=False)
    monkeypatch.setenv("APEX_PROJECT_MINER_PLANNING_ROOT", str(planning_root))
    clear_seed_cache()

    seed = load_seed_data()

    assert seed["metadata"]["equipment_workbook_path"] == str(equipment_path)
    assert seed["metadata"]["capability_workbook_path"] == str(capability_path)
    assert seed["crew"][0]["name"] == "Ace Randolph"
    assert seed["equipment_inventory"][0]["assigned_to_id"] == "tech-001"
