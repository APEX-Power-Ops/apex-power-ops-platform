from __future__ import annotations

import os
import re
from collections import Counter
from functools import lru_cache
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional


DEFAULT_PROJECT_MINER_PLANNING_ROOT = Path.home() / "Desktop" / "Project Miner PM Planning"
DEFAULT_ESTIMATOR_WORKBOOK_NAME = "Estimator R3 - Project Miner Temp Power Testing.xlsm"
DEFAULT_SLD_PDF_NAME = "Miner Temp SLD-AP-BCARRASCO.pdf"
LEGACY_ESTIMATOR_WORKBOOK_PATH = Path.home() / "Desktop" / DEFAULT_ESTIMATOR_WORKBOOK_NAME
LEGACY_SLD_PDF_PATH = Path.home() / "Desktop" / DEFAULT_SLD_PDF_NAME
FLAT_ESTIMATOR_SHEET_NAMES = ("Updated", "Quote Tab")
SCOPE_REFERENCE_SHEET_NAME = "Equipment Reference"
SCOPE_REFERENCE_START_ROW = 4
SCOPE_REFERENCE_END_ROW = 80
APPARATUS_START_ROW = 6
APPARATUS_END_ROW = 488

TOPOLOGY_PATTERNS = [
    re.compile(r"PWR\s+SKID\s*-\s*[A-Z0-9]+", re.IGNORECASE),
    re.compile(r"MVTX\s*-\s*[A-Z0-9]+", re.IGNORECASE),
    re.compile(r"SWBD\s*-\s*[A-Z0-9]+", re.IGNORECASE),
    re.compile(r"MVS\s*-\s*[A-Z0-9]+", re.IGNORECASE),
    re.compile(r"LVPP\s*-\s*[A-Z0-9]+", re.IGNORECASE),
    re.compile(r"LVTX\s*-\s*[A-Z0-9]+", re.IGNORECASE),
]


def _clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    return value


def _normalize_label(label: str) -> str:
    return re.sub(r"\s+", " ", label.strip()).upper()


def _as_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except Exception:
        return None


def _as_positive_int(value: Any) -> Optional[int]:
    number = _as_float(value)
    if number is None or number <= 0:
        return None
    return int(number)


def _first_existing_path(*paths: Path) -> Path:
    for path in paths:
        if path.exists():
            return path
    return paths[0]


def _project_miner_planning_root() -> Path:
    return Path(
        os.getenv("APEX_PROJECT_MINER_PLANNING_ROOT", str(DEFAULT_PROJECT_MINER_PLANNING_ROOT)).strip()
    )


def _default_paths() -> tuple[str, str]:
    planning_root = _project_miner_planning_root()
    workbook_path = os.getenv("APEX_PROJECT_ESTIMATOR_WORKBOOK")
    sld_path = os.getenv("APEX_PROJECT_SLD_PDF")
    resolved_workbook_path = Path(workbook_path.strip()) if workbook_path else _first_existing_path(
        planning_root / DEFAULT_ESTIMATOR_WORKBOOK_NAME,
        LEGACY_ESTIMATOR_WORKBOOK_PATH,
    )
    resolved_sld_path = Path(sld_path.strip()) if sld_path else _first_existing_path(
        planning_root / DEFAULT_SLD_PDF_NAME,
        LEGACY_SLD_PDF_PATH,
    )
    return str(resolved_workbook_path), str(resolved_sld_path)


def clear_project_seed_cache() -> None:
    load_project_seed_sources.cache_clear()


def extract_topology_labels_from_text(text: str) -> List[str]:
    labels: List[str] = []
    seen = set()
    for pattern in TOPOLOGY_PATTERNS:
        for match in pattern.finditer(text or ""):
            normalized = _normalize_label(match.group(0))
            if normalized not in seen:
                seen.add(normalized)
                labels.append(normalized)
    return labels


def _load_pdf_topology_labels(pdf_path: Path) -> Dict[str, Any]:
    if not pdf_path.exists():
        return {
            "pdf_found": False,
            "topology_labels": [],
            "topology_counts": {},
            "pdf_pages_read": 0,
        }

    try:
        from pypdf import PdfReader
    except Exception:
        return {
            "pdf_found": True,
            "topology_labels": [],
            "topology_counts": {},
            "pdf_pages_read": 0,
        }

    reader = PdfReader(str(pdf_path))
    labels: List[str] = []
    seen = set()
    pages_read = 0
    for page in reader.pages[:8]:
        pages_read += 1
        text = page.extract_text() or ""
        for label in extract_topology_labels_from_text(text):
            if label not in seen:
                seen.add(label)
                labels.append(label)

    counts = Counter(label.split("-")[0].strip() for label in labels)
    return {
        "pdf_found": True,
        "topology_labels": labels,
        "topology_counts": dict(counts),
        "pdf_pages_read": pages_read,
    }


def _empty_workbook_data(found: bool) -> Dict[str, Any]:
    return {
        "workbook_found": found,
        "sheet_name": None,
        "source_format": None,
        "scope_sheets": [],
        "scope_count": 0,
        "project_name": None,
        "location": None,
        "drawing_package": None,
        "issue_date": None,
        "line_items": [],
        "expanded_apparatus_candidates": [],
    }


def _append_expanded_candidates(
    expanded_candidates: List[Dict[str, Any]],
    line_item: Dict[str, Any],
    apparatus_counter: int,
) -> int:
    quantity = max(1, int(line_item.get("qty") or 1))
    for item_index in range(1, quantity + 1):
        apparatus_counter += 1
        suffix = f"{item_index:02d}" if quantity > 1 else "01"
        display_name = line_item.get("designation") or line_item.get("apparatus_type")
        expanded_candidates.append(
            {
                "candidate_id": f"miner-app-{apparatus_counter:04d}",
                "line_id": line_item["line_id"],
                "section": line_item.get("section"),
                "apparatus_type": line_item.get("apparatus_type"),
                "designation": line_item.get("designation"),
                "drawing_ref": line_item.get("drawing_ref"),
                "display_name": f"{display_name} {suffix}" if quantity > 1 else str(display_name),
                "planned_hours": line_item.get("hrs_per_unit"),
                "source_row": line_item.get("source_row"),
                "scope_sheet": line_item.get("scope_sheet"),
            }
        )
    return apparatus_counter


def _flat_sheet_line_items(sheet_name: str, sheet: Any) -> Dict[str, Any]:
    row3 = [_clean(value) for value in next(sheet.iter_rows(min_row=3, max_row=3, values_only=True))]
    row4 = [_clean(value) for value in next(sheet.iter_rows(min_row=4, max_row=4, values_only=True))]

    line_items: List[Dict[str, Any]] = []
    expanded_candidates: List[Dict[str, Any]] = []
    line_index = 0
    apparatus_counter = 0

    for row_number, row in enumerate(sheet.iter_rows(min_row=6, values_only=True), start=6):
        values = [_clean(value) for value in row]
        qty = values[2] if len(values) > 2 else None
        section = values[3] if len(values) > 3 else None
        apparatus_type = values[4] if len(values) > 4 else None
        designation = values[5] if len(values) > 5 else None
        notes = values[6] if len(values) > 6 else None
        hrs_per_unit = values[8] if len(values) > 8 else None
        hrs_line = values[9] if len(values) > 9 else None

        if apparatus_type is None:
            continue

        quantity = _as_positive_int(qty) or 1
        line_index += 1
        line_item = {
            "line_id": f"miner-line-{line_index:03d}",
            "qty": quantity,
            "section": section,
            "apparatus_type": apparatus_type,
            "designation": designation,
            "drawing_ref": notes,
            "hrs_per_unit": hrs_per_unit,
            "hrs_line": hrs_line,
            "source_row": row_number,
        }
        line_items.append(line_item)
        apparatus_counter = _append_expanded_candidates(expanded_candidates, line_item, apparatus_counter)

    return {
        "workbook_found": True,
        "sheet_name": sheet_name,
        "source_format": "flat_quote",
        "scope_sheets": [],
        "scope_count": 0,
        "project_name": row3[4] if len(row3) > 4 else None,
        "location": row4[4] if len(row4) > 4 else None,
        "drawing_package": row3[5] if len(row3) > 5 else None,
        "issue_date": row4[5] if len(row4) > 5 else None,
        "line_items": line_items,
        "expanded_apparatus_candidates": expanded_candidates,
    }


def _scope_sheet_names(workbook: Any) -> List[str]:
    if SCOPE_REFERENCE_SHEET_NAME not in workbook.sheetnames:
        return []

    reference_sheet = workbook[SCOPE_REFERENCE_SHEET_NAME]
    scope_names: List[str] = []
    for row in reference_sheet.iter_rows(
        min_row=SCOPE_REFERENCE_START_ROW,
        max_row=SCOPE_REFERENCE_END_ROW,
        min_col=12,
        max_col=13,
        values_only=True,
    ):
        sheet_name = _clean(row[0])
        scope_total = _as_float(row[1])
        if not sheet_name or sheet_name not in workbook.sheetnames:
            continue
        if scope_total is None or scope_total <= 0:
            continue
        if sheet_name in FLAT_ESTIMATOR_SHEET_NAMES:
            continue
        scope_names.append(str(sheet_name))
    return scope_names


def _scope_sheet_line_items(workbook_path: Path, workbook: Any) -> Dict[str, Any]:
    scope_names = _scope_sheet_names(workbook)
    line_items: List[Dict[str, Any]] = []
    expanded_candidates: List[Dict[str, Any]] = []
    line_index = 0
    apparatus_counter = 0

    for scope_name in scope_names:
        sheet = workbook[scope_name]
        scope_type = _clean(sheet["C4"].value)
        scope_total_hours = sheet["J3"].value
        scope_multiplier = sheet["M4"].value
        scope_quoted_amount = sheet["P3"].value
        current_section = scope_name

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=APPARATUS_START_ROW, max_row=APPARATUS_END_ROW, values_only=True),
            start=APPARATUS_START_ROW,
        ):
            values = [_clean(value) for value in row]
            qty = values[2] if len(values) > 2 else None
            apparatus_type = values[4] if len(values) > 4 else None
            hrs_per_unit = values[8] if len(values) > 8 else None
            hrs_line = values[9] if len(values) > 9 else None

            if apparatus_type is None:
                continue

            quantity = _as_positive_int(qty)
            if quantity is None:
                current_section = str(apparatus_type)
                continue

            line_index += 1
            line_item = {
                "line_id": f"miner-line-{line_index:03d}",
                "qty": quantity,
                "section": current_section,
                "apparatus_type": apparatus_type,
                "designation": None,
                "drawing_ref": scope_name,
                "hrs_per_unit": hrs_per_unit,
                "hrs_line": hrs_line,
                "source_row": row_number,
                "scope_sheet": scope_name,
                "scope_type": scope_type,
                "scope_total_hours": scope_total_hours,
                "scope_multiplier": scope_multiplier,
                "scope_quoted_amount": scope_quoted_amount,
            }
            line_items.append(line_item)
            apparatus_counter = _append_expanded_candidates(expanded_candidates, line_item, apparatus_counter)

    return {
        "workbook_found": True,
        "sheet_name": "scope_sheets" if scope_names else None,
        "source_format": "scope_sheets" if scope_names else None,
        "scope_sheets": scope_names,
        "scope_count": len(scope_names),
        "project_name": workbook_path.stem if scope_names else None,
        "location": None,
        "drawing_package": None,
        "issue_date": None,
        "line_items": line_items,
        "expanded_apparatus_candidates": expanded_candidates,
    }


def _workbook_line_items(workbook_path: Path) -> Dict[str, Any]:
    if not workbook_path.exists():
        return _empty_workbook_data(False)

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(workbook_path.read_bytes()), data_only=True, read_only=True)
    try:
        for sheet_name in FLAT_ESTIMATOR_SHEET_NAMES:
            if sheet_name not in workbook.sheetnames:
                continue
            flat_data = _flat_sheet_line_items(sheet_name, workbook[sheet_name])
            if flat_data["line_items"]:
                return flat_data

        scope_data = _scope_sheet_line_items(workbook_path, workbook)
        if scope_data["line_items"]:
            return scope_data
    finally:
        workbook.close()

    return _empty_workbook_data(True)


@lru_cache(maxsize=4)
def load_project_seed_sources(
    estimator_workbook_path: Optional[str] = None,
    sld_pdf_path: Optional[str] = None,
) -> Dict[str, Any]:
    workbook_path = Path(estimator_workbook_path or _default_paths()[0])
    pdf_path = Path(sld_pdf_path or _default_paths()[1])

    workbook_data = _workbook_line_items(workbook_path)
    pdf_data = _load_pdf_topology_labels(pdf_path)

    return {
        "metadata": {
            "estimator_workbook_path": str(workbook_path),
            "estimator_workbook_found": workbook_data["workbook_found"],
            "sld_pdf_path": str(pdf_path),
            "sld_pdf_found": pdf_data["pdf_found"],
            "estimator_source_format": workbook_data["source_format"],
            "estimator_scope_count": workbook_data["scope_count"],
        },
        "project_name": workbook_data["project_name"],
        "location": workbook_data["location"],
        "drawing_package": workbook_data["drawing_package"],
        "issue_date": workbook_data["issue_date"],
        "source_sheet": workbook_data["sheet_name"],
        "source_format": workbook_data["source_format"],
        "scope_sheets": workbook_data["scope_sheets"],
        "scope_count": workbook_data["scope_count"],
        "line_items": workbook_data["line_items"],
        "expanded_apparatus_candidates": workbook_data["expanded_apparatus_candidates"],
        "topology_labels": pdf_data["topology_labels"],
        "topology_counts": pdf_data["topology_counts"],
        "pdf_pages_read": pdf_data["pdf_pages_read"],
    }
