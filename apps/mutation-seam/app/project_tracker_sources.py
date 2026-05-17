from __future__ import annotations

import os
from collections import Counter
from functools import lru_cache
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional


DEFAULT_PROJECT_MINER_PLANNING_ROOT = Path.home() / "Desktop" / "Project Miner PM Planning"
DEFAULT_PROJECT_DATA_ENTRY_WORKBOOK_NAME = "RESA Power - Project Data Entry MASTER.xlsm"
DEFAULT_REFERENCE_TRACKER_WORKBOOK_NAME = "Garney- Central Mesa Reuse Tracker #677562.xlsm"
PROJECT_FORM_SHEET_NAME = "Project_Form"
TASK_ENTRY_SHEET_NAME = "Task_Entry"
ALL_TASKS_SHEET_NAME = "All_Tasks"
PROJECT_FORM_METADATA_ROWS = range(4, 24)
PROJECT_FORM_WORKSCOPE_ROWS = range(4, 24)
TASK_ENTRY_COLUMNS = [
    ("scope", "Scope"),
    ("neta_standard", "NETA_Standard"),
    ("task_id", "Task_ID"),
    ("task", "Task"),
    ("apparatus", "Apparatus"),
    ("designation", "Designation"),
    ("drawing", "Drawing"),
    ("apparatus_hours", "Apparatus_Hourrs"),
]
ALL_TASK_COLUMNS = [
    ("scope", "Scope"),
    ("neta_standard", "NETA_Standard"),
    ("task_id", "Task_ID"),
    ("task", "Task"),
    ("apparatus", "Apparatus"),
    ("designation", "Designation"),
    ("drawing", "Drawing"),
    ("date_due", "Date Due"),
    ("notes", "Notes"),
    ("assessment", "Assessment"),
    ("datasheet", "DATASHEET"),
    ("date_completed", "DATE COMPLETED"),
    ("notes2", "NOTES2"),
    ("completion", "% COMPLETION"),
    ("task_delays", "TASK DELAYS"),
    ("apparatus_hours", "Apparatus Hours"),
    ("remaining_hours", "Remaining Hours"),
    ("actual_hours", "ACTUAL HOURS"),
    ("status", "STATUS"),
    ("availability", "AVAILABILITY"),
    ("priority", "PRIORITY"),
    ("apparatus_category", "Apparatus Category"),
]
FORMULA_ERROR_SAMPLE_LIMIT = 5
FORMULA_ERROR_SAMPLE_KEYS = ["scope", "task_id", "task", "apparatus", "designation"]


def _clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    return value


def _project_miner_planning_root() -> Path:
    return Path(
        os.getenv("APEX_PROJECT_MINER_PLANNING_ROOT", str(DEFAULT_PROJECT_MINER_PLANNING_ROOT)).strip()
    )


def _workbook_paths() -> tuple[str, str]:
    planning_root = _project_miner_planning_root()
    data_entry_path = os.getenv("APEX_PROJECT_DATA_ENTRY_WORKBOOK")
    tracker_path = os.getenv("APEX_REFERENCE_TRACKER_WORKBOOK")
    resolved_data_entry_path = (
        Path(data_entry_path.strip())
        if data_entry_path
        else planning_root / DEFAULT_PROJECT_DATA_ENTRY_WORKBOOK_NAME
    )
    resolved_tracker_path = (
        Path(tracker_path.strip())
        if tracker_path
        else planning_root / DEFAULT_REFERENCE_TRACKER_WORKBOOK_NAME
    )
    return str(resolved_data_entry_path), str(resolved_tracker_path)


def clear_project_tracker_cache() -> None:
    load_project_tracker_sources.cache_clear()


def _project_form_data(workbook: Any) -> Dict[str, Any]:
    if PROJECT_FORM_SHEET_NAME not in workbook.sheetnames:
        return {
            "metadata": {},
            "workscope_names": [],
        }

    sheet = workbook[PROJECT_FORM_SHEET_NAME]
    metadata: Dict[str, Any] = {}
    for row_number in PROJECT_FORM_METADATA_ROWS:
        label = _clean_cell(sheet[f"B{row_number}"].value)
        value = _clean_cell(sheet[f"C{row_number}"].value)
        if label is None:
            continue
        metadata[str(label).rstrip(":")] = value

    workscope_names = [
        str(value)
        for row_number in PROJECT_FORM_WORKSCOPE_ROWS
        if (value := _clean_cell(sheet[f"E{row_number}"].value)) is not None
    ]
    return {
        "metadata": metadata,
        "workscope_names": workscope_names,
    }


def _row_has_data(row: Dict[str, Any], keys: List[str]) -> bool:
    return any(row.get(key) is not None for key in keys)


def _read_rows(
    sheet: Any,
    columns: List[tuple[str, str]],
    identity_keys: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    keys = identity_keys or [key for key, _ in columns]
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        parsed = {"source_row": row_number}
        for index, (key, _) in enumerate(columns):
            parsed[key] = _clean_cell(values[index]) if index < len(values) else None
        if _row_has_data(parsed, keys):
            rows.append(parsed)
    return rows


def _formula_error_summary(rows: List[Dict[str, Any]], columns: List[tuple[str, str]]) -> Dict[str, Any]:
    label_by_key = {key: label for key, label in columns}
    column_counts: Counter[str] = Counter()
    sample_rows: List[Dict[str, Any]] = []
    cell_count = 0
    row_count = 0
    for row in rows:
        error_columns: List[str] = []
        for key, value in row.items():
            if isinstance(value, str) and value.startswith("#"):
                cell_count += 1
                column_label = label_by_key.get(key, key)
                column_counts[column_label] += 1
                error_columns.append(column_label)
        if error_columns:
            row_count += 1
            if len(sample_rows) < FORMULA_ERROR_SAMPLE_LIMIT:
                sample = {"source_row": row.get("source_row"), "error_columns": error_columns}
                sample.update({key: row.get(key) for key in FORMULA_ERROR_SAMPLE_KEYS if row.get(key) is not None})
                sample_rows.append(sample)
    return {
        "formula_error_cell_count": cell_count,
        "formula_error_row_count": row_count,
        "formula_error_column_counts": dict(column_counts),
        "formula_error_sample_rows": sample_rows,
    }


def _counter(values: List[Any]) -> Dict[str, int]:
    return {
        str(key): count
        for key, count in Counter(value for value in values if value is not None).items()
    }


def _read_planning_workbook(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {
            "path": str(path),
            "found": False,
            "project_form": {},
            "workscope_names": [],
            "task_entry_count": 0,
            "all_tasks_count": 0,
            "task_entry_scope_counts": {},
            "all_tasks_scope_counts": {},
            "status_counts": {},
            "availability_counts": {},
            "apparatus_category_counts": {},
            "formula_error_cell_count": 0,
            "formula_error_row_count": 0,
            "formula_error_column_counts": {},
            "formula_error_sample_rows": [],
            "task_entry_sample": [],
            "all_tasks_sample": [],
        }

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(path.read_bytes()), data_only=True, read_only=True)
    try:
        project_form = _project_form_data(workbook)
        task_entry_rows = (
            _read_rows(workbook[TASK_ENTRY_SHEET_NAME], TASK_ENTRY_COLUMNS)
            if TASK_ENTRY_SHEET_NAME in workbook.sheetnames
            else []
        )
        all_tasks_rows = (
            _read_rows(
                workbook[ALL_TASKS_SHEET_NAME],
                ALL_TASK_COLUMNS,
                identity_keys=["scope", "task_id", "task", "apparatus"],
            )
            if ALL_TASKS_SHEET_NAME in workbook.sheetnames
            else []
        )
        error_counts = _formula_error_summary(all_tasks_rows, ALL_TASK_COLUMNS)
        return {
            "path": str(path),
            "found": True,
            "project_form": project_form["metadata"],
            "workscope_names": project_form["workscope_names"],
            "task_entry_count": len(task_entry_rows),
            "all_tasks_count": len(all_tasks_rows),
            "task_entry_scope_counts": _counter([row.get("scope") for row in task_entry_rows]),
            "all_tasks_scope_counts": _counter([row.get("scope") for row in all_tasks_rows]),
            "status_counts": _counter([row.get("status") for row in all_tasks_rows]),
            "availability_counts": _counter([row.get("availability") for row in all_tasks_rows]),
            "apparatus_category_counts": _counter([row.get("apparatus_category") for row in all_tasks_rows]),
            **error_counts,
            "task_entry_sample": task_entry_rows[:5],
            "all_tasks_sample": all_tasks_rows[:5],
        }
    finally:
        workbook.close()


@lru_cache(maxsize=4)
def load_project_tracker_sources(
    project_data_entry_workbook_path: Optional[str] = None,
    reference_tracker_workbook_path: Optional[str] = None,
) -> Dict[str, Any]:
    default_data_entry_path, default_tracker_path = _workbook_paths()
    data_entry_path = Path(project_data_entry_workbook_path or default_data_entry_path)
    tracker_path = Path(reference_tracker_workbook_path or default_tracker_path)

    data_entry = _read_planning_workbook(data_entry_path)
    reference_tracker = _read_planning_workbook(tracker_path)
    return {
        "metadata": {
            "project_data_entry_workbook_path": str(data_entry_path),
            "project_data_entry_workbook_found": data_entry["found"],
            "reference_tracker_workbook_path": str(tracker_path),
            "reference_tracker_workbook_found": reference_tracker["found"],
        },
        "project_data_entry": data_entry,
        "reference_tracker": reference_tracker,
    }
