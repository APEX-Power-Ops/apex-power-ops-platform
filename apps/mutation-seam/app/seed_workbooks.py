from __future__ import annotations

import os
import re
from functools import lru_cache
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional


DEFAULT_PROJECT_MINER_PLANNING_ROOT = Path.home() / "Desktop" / "Project Miner PM Planning"
DEFAULT_EQUIPMENT_WORKBOOK_NAME = "EQUIPMENT INVENTORY - 2026.xlsx"
DEFAULT_CAPABILITY_WORKBOOK_NAME = "Phx Tech Testing Capability Matrix 032726.xlsx"
LEGACY_EQUIPMENT_WORKBOOK_PATH = Path.home() / "Desktop" / DEFAULT_EQUIPMENT_WORKBOOK_NAME
LEGACY_CAPABILITY_WORKBOOK_PATH = Path.home() / "Desktop" / DEFAULT_CAPABILITY_WORKBOOK_NAME

FALLBACK_CREW = [
    {"id": "tech-001", "name": "Alex Rivera", "role": "field_tech"},
    {"id": "tech-002", "name": "Sam Chen", "role": "field_tech"},
    {"id": "tech-003", "name": "Jordan Bell", "role": "field_tech"},
]


def _normalize_name(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\s*\([^)]*\)\s*$", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.casefold()


def _clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    return value


def _first_existing_path(*paths: Path) -> Path:
    for path in paths:
        if path.exists():
            return path
    return paths[0]


def _project_miner_planning_root() -> Path:
    return Path(
        os.getenv("APEX_PROJECT_MINER_PLANNING_ROOT", str(DEFAULT_PROJECT_MINER_PLANNING_ROOT)).strip()
    )


def _workbook_paths() -> tuple[str, str]:
    planning_root = _project_miner_planning_root()
    equipment_path = os.getenv("APEX_FIELD_SEED_EQUIPMENT_WORKBOOK")
    capability_path = os.getenv("APEX_FIELD_SEED_CAPABILITY_WORKBOOK")
    resolved_equipment_path = Path(equipment_path.strip()) if equipment_path else _first_existing_path(
        planning_root / DEFAULT_EQUIPMENT_WORKBOOK_NAME,
        LEGACY_EQUIPMENT_WORKBOOK_PATH,
    )
    resolved_capability_path = Path(capability_path.strip()) if capability_path else _first_existing_path(
        planning_root / DEFAULT_CAPABILITY_WORKBOOK_NAME,
        LEGACY_CAPABILITY_WORKBOOK_PATH,
    )
    return str(resolved_equipment_path), str(resolved_capability_path)


def clear_seed_cache() -> None:
    load_seed_data.cache_clear()


@lru_cache(maxsize=4)
def load_seed_data(
    equipment_workbook_path: Optional[str] = None,
    capability_workbook_path: Optional[str] = None,
) -> Dict[str, Any]:
    equipment_path = Path(equipment_workbook_path or _workbook_paths()[0])
    capability_path = Path(capability_workbook_path or _workbook_paths()[1])

    try:
        from openpyxl import load_workbook
    except Exception:
        return {
            "metadata": {
                "seed_source": "fallback",
                "equipment_workbook_found": False,
                "capability_workbook_found": False,
            },
            "crew": FALLBACK_CREW,
            "equipment_inventory": [],
            "standard_tech_list": [],
            "tech_capabilities": [],
            "capability_score_scale": [],
        }

    crew = list(FALLBACK_CREW)
    equipment_inventory: List[Dict[str, Any]] = []
    standard_tech_list: List[Dict[str, Any]] = []
    tech_capabilities: List[Dict[str, Any]] = []
    capability_score_scale: List[Dict[str, Any]] = []
    crew_lookup: Dict[str, Dict[str, Any]] = {_normalize_name(item["name"]): item for item in crew}

    if capability_path.exists():
        workbook = load_workbook(BytesIO(capability_path.read_bytes()), data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            total_row = list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))
            name_row = list(next(sheet.iter_rows(min_row=3, max_row=3, values_only=True)))

            parsed_crew: List[Dict[str, Any]] = []
            for index, raw_name in enumerate(name_row[3:], start=4):
                cleaned_name = _clean_cell(raw_name)
                if cleaned_name is None:
                    continue
                tech_id = f"tech-{len(parsed_crew) + 1:03d}"
                parsed_crew.append(
                    {
                        "id": tech_id,
                        "name": re.sub(r"\s*\([^)]*\)\s*$", "", str(cleaned_name)).strip(),
                        "role": "field_tech",
                        "experience_score": total_row[index - 1],
                        "source_column_label": cleaned_name,
                    }
                )

            if parsed_crew:
                crew = parsed_crew
                crew_lookup = {_normalize_name(item["name"]): item for item in crew}

            scale_lookup: Dict[Any, Any] = {}
            for row in sheet.iter_rows(min_row=4, values_only=True):
                row_values = [_clean_cell(value) for value in row]
                if row_values[0] is None and row_values[1] is None and row_values[2] is None:
                    continue
                if row_values[22] is not None and row_values[23] is not None:
                    scale_lookup[row_values[22]] = row_values[23]
                scores = {}
                for crew_index, crew_item in enumerate(crew, start=4):
                    score_value = row_values[crew_index - 1] if crew_index - 1 < len(row_values) else None
                    if score_value is not None:
                        scores[crew_item["id"]] = score_value
                tech_capabilities.append(
                    {
                        "type": row_values[0],
                        "voltage": row_values[1],
                        "test": row_values[2],
                        "scores": scores,
                    }
                )
            capability_score_scale = [
                {"score": score, "label": label} for score, label in sorted(scale_lookup.items(), key=lambda item: item[0])
            ]
        finally:
            workbook.close()

    if equipment_path.exists():
        workbook = load_workbook(BytesIO(equipment_path.read_bytes()), data_only=True)
        try:
            if "ALL Equipment" in workbook.sheetnames:
                sheet = workbook["ALL Equipment"]
                for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                    row_values = [_clean_cell(value) for value in row]
                    if not any(row_values[:9]):
                        continue
                    assigned_name = row_values[8]
                    assigned_tech = crew_lookup.get(_normalize_name(assigned_name)) if assigned_name else None
                    equipment_inventory.append(
                        {
                            "id": f"equip-{row_number:04d}",
                            "category": row_values[0],
                            "inventory_id": row_values[1],
                            "equipment": row_values[2],
                            "manufacturer": row_values[3],
                            "model": row_values[4],
                            "serial_number": row_values[5],
                            "last_cal_date": row_values[6],
                            "cal_date_2026": row_values[7],
                            "assigned_to_name": assigned_name,
                            "assigned_to_id": assigned_tech["id"] if assigned_tech else None,
                            "notes": row_values[9],
                        }
                    )

            if "Standard Tech List" in workbook.sheetnames:
                sheet = workbook["Standard Tech List"]
                header_row = [_clean_cell(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
                tech_columns = []
                for column_index, raw_name in enumerate(header_row[4:], start=5):
                    cleaned_name = _clean_cell(raw_name)
                    if cleaned_name is None or str(cleaned_name).strip().lower() == "name":
                        continue
                    tech_item = crew_lookup.get(_normalize_name(cleaned_name))
                    tech_columns.append(
                        {
                            "column_index": column_index,
                            "name": cleaned_name,
                            "tech_id": tech_item["id"] if tech_item else None,
                        }
                    )
                for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                    row_values = [_clean_cell(value) for value in row]
                    if row_values[0] is None and row_values[2] is None:
                        continue
                    assignments = {}
                    for column in tech_columns:
                        raw_value = row_values[column["column_index"] - 1] if column["column_index"] - 1 < len(row_values) else None
                        if raw_value is not None:
                            key = column["tech_id"] or column["name"]
                            assignments[key] = raw_value
                    standard_tech_list.append(
                        {
                            "id": f"std-equip-{row_number:03d}",
                            "category": row_values[0],
                            "phoenix_id": row_values[1],
                            "equipment": row_values[2],
                            "standard_optional": row_values[3],
                            "technician_assignments": assignments,
                        }
                    )
        finally:
            workbook.close()

    return {
        "metadata": {
            "seed_source": "workbooks" if capability_path.exists() or equipment_path.exists() else "fallback",
            "equipment_workbook_path": str(equipment_path),
            "equipment_workbook_found": equipment_path.exists(),
            "capability_workbook_path": str(capability_path),
            "capability_workbook_found": capability_path.exists(),
        },
        "crew": crew,
        "equipment_inventory": equipment_inventory,
        "standard_tech_list": standard_tech_list,
        "tech_capabilities": tech_capabilities,
        "capability_score_scale": capability_score_scale,
    }
