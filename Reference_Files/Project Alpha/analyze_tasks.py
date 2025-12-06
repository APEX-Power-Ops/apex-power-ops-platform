import openpyxl
from collections import Counter

output_file = r'C:\RESA_Power_Build\Reference_Files\Project Alpha\task_analysis.txt'
xlsm_path = r'C:\RESA_Power_Build\Reference_Files\Project Alpha\Delta Diversified - Alpha LV Tracker.xlsm'

with open(output_file, 'w', encoding='utf-8') as f:
    wb = openpyxl.load_workbook(xlsm_path, data_only=True)
    ws = wb['All_Tasks']
    
    f.write("=" * 70 + "\n")
    f.write("ALL_TASKS ANALYSIS - TASK COLUMN VALUES\n")
    f.write("=" * 70 + "\n\n")
    
    # Collect all Task values (column 4)
    tasks = []
    designations = []
    drawings = []
    scopes = []
    
    for row_num in range(2, ws.max_row + 1):
        task = ws.cell(row=row_num, column=4).value
        designation = ws.cell(row=row_num, column=6).value
        drawing = ws.cell(row=row_num, column=7).value
        scope = ws.cell(row=row_num, column=1).value
        
        if task:
            tasks.append(str(task))
        if designation:
            designations.append(str(designation))
        if drawing:
            drawings.append(str(drawing))
        if scope:
            scopes.append(str(scope))
    
    f.write(f"Total Rows: {ws.max_row - 1}\n")
    f.write(f"Tasks with values: {len(tasks)}\n")
    f.write(f"Designations filled: {len(designations)}\n")
    f.write(f"Drawings filled: {len(drawings)}\n\n")
    
    # Unique task values
    task_counts = Counter(tasks)
    f.write(f"Unique Task Values: {len(task_counts)}\n\n")
    
    f.write("TOP 50 TASK VALUES (by frequency):\n")
    f.write("-" * 50 + "\n")
    for task, count in task_counts.most_common(50):
        f.write(f"{count:4d}x  {task}\n")
    
    # Scope breakdown
    scope_counts = Counter(scopes)
    f.write("\n\nSCOPE BREAKDOWN:\n")
    f.write("-" * 50 + "\n")
    for scope, count in scope_counts.most_common():
        f.write(f"{count:4d}x  {scope}\n")
    
    # Sample rows with SLD-like task names
    f.write("\n\nSAMPLE ROWS WITH 1F0 PREFIX IN TASK:\n")
    f.write("-" * 70 + "\n")
    count = 0
    for row_num in range(2, ws.max_row + 1):
        task = ws.cell(row=row_num, column=4).value
        if task and '1F0' in str(task):
            scope = ws.cell(row=row_num, column=1).value
            apparatus = ws.cell(row=row_num, column=5).value
            designation = ws.cell(row=row_num, column=6).value
            drawing = ws.cell(row=row_num, column=7).value
            f.write(f"\nRow {row_num}:\n")
            f.write(f"  Scope: {scope}\n")
            f.write(f"  Task: {task}\n")
            f.write(f"  Apparatus: {apparatus}\n")
            f.write(f"  Designation: {designation}\n")
            f.write(f"  Drawing: {drawing}\n")
            count += 1
            if count >= 20:
                break
    
    wb.close()
    f.write("\n\nAnalysis complete.\n")

print("Done")
