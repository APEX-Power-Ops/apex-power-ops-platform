import openpyxl

output_file = r'C:\RESA_Power_Build\Reference_Files\Project Alpha\xlsm_analysis.txt'
xlsm_path = r'C:\RESA_Power_Build\Reference_Files\Project Alpha\Delta Diversified - Alpha LV Tracker.xlsm'

with open(output_file, 'w', encoding='utf-8') as f:
    try:
        # Not using read_only mode to get full access
        wb = openpyxl.load_workbook(xlsm_path, data_only=True)
        f.write("=" * 70 + "\n")
        f.write("DELTA DIVERSIFIED - ALPHA LV TRACKER.XLSM ANALYSIS\n")
        f.write("=" * 70 + "\n\n")
        
        f.write(f"Total Sheets: {len(wb.sheetnames)}\n\n")
        
        for i, sheet_name in enumerate(wb.sheetnames):
            f.write(f"{i+1}. {sheet_name}\n")
        
        # Focus on key sheets
        key_sheets = ['Apparatus_List_w_Hours', 'All_Tasks', 'E1 Electrical Room', 'E2 Electrical Room']
        
        f.write("\n" + "=" * 70 + "\n")
        f.write("KEY SHEET DETAILS\n")
        f.write("=" * 70 + "\n")
        
        for sheet_name in key_sheets:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                f.write(f"\n{'='*50}\n")
                f.write(f"SHEET: {sheet_name}\n")
                f.write(f"{'='*50}\n")
                f.write(f"Max Row: {ws.max_row}, Max Column: {ws.max_column}\n\n")
                
                # Get headers from first row
                f.write("COLUMNS:\n")
                headers = []
                for col_num, cell in enumerate(ws[1], 1):
                    val = str(cell.value) if cell.value else "(empty)"
                    headers.append(val)
                    f.write(f"  Col {col_num}: {val}\n")
                
                # Sample data rows 2-6
                f.write("\nSAMPLE DATA (rows 2-6):\n")
                for row_num in range(2, min(7, ws.max_row + 1)):
                    f.write(f"\n  Row {row_num}:\n")
                    for col_num, cell in enumerate(ws[row_num], 1):
                        if cell.value is not None:
                            header = headers[col_num-1] if col_num <= len(headers) else f"Col{col_num}"
                            val = str(cell.value)[:50]
                            f.write(f"    {header}: {val}\n")
        
        wb.close()
        f.write("\n\nAnalysis complete.\n")
        
    except Exception as e:
        f.write(f"\nError: {str(e)}\n")
        import traceback
        f.write(traceback.format_exc())

print("Done")
