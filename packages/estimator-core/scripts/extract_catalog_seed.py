#!/usr/bin/env python3
"""Extract the apparatus catalog from the master workbook's Equipment Reference sheet.

Usage:
  uv run --with openpyxl python scripts/extract_catalog_seed.py \
      "G:/My Drive/APEX Platform/Estimator PHX 012326 (Master).xlsm" \
      src/catalog/equipment-models.seed.json

Reads Equipment Reference cols B (ATS section), C (MTS section), D (Scope of Work / apparatus),
E (ATS25 hrs), F (MTS23 hrs) for data rows (row 3..end). Emits one EquipmentModel per apparatus.
'NA' (any case) -> null hours; trailing '(Set)' -> unit_of_issue 'set'. All rows seed as active.
"""
import json
import sys
from openpyxl import load_workbook


def parse_hours(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == '' or s.upper() == 'NA':
        return None
    return float(s)


def main(xlsx_path: str, out_path: str) -> None:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb['Equipment Reference']
    rows = []
    seen = set()
    for row in ws.iter_rows(min_row=3, min_col=2, max_col=6, values_only=True):
        ats_section, mts_section, apparatus, ats25, mts23 = row
        if apparatus is None or str(apparatus).strip() == '':
            continue
        ref = str(apparatus).strip()
        if ref in seen:  # first occurrence wins
            continue
        seen.add(ref)
        unit = 'set' if ref.endswith('(Set)') else 'each'
        rows.append({
            'ref': ref,
            'apparatus': ref,
            'neta_section': {
                'ATS': (str(ats_section).strip() if ats_section is not None else None),
                'MTS': (str(mts_section).strip() if mts_section is not None else None),
            },
            'ref_hours': {'ATS': parse_hours(ats25), 'MTS': parse_hours(mts23)},
            'unit_of_issue': unit,
            'lifecycle_status': 'active',
            'merged_into_ref': None,
        })
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)
        f.write('\n')
    print(f'wrote {len(rows)} apparatus rows -> {out_path}')


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])
