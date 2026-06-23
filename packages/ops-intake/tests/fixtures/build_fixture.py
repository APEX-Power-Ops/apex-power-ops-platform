"""Builds a tiny estimator workbook mirroring the real cell map, for fast extractor tests.

Numbers are chosen to reconcile exactly:
  one scope, J3 = 25 hrs (lines 2x5 + 3x5), categories 800/100/50/50 = P3 1000, M4=N4=1 => P4 1000,
  Print_Template TOTAL COST 1000.

Task 4 additions:
  - row 6: bold section header above apparatus rows
  - N4: explicit 1.0 (pct-adjust)
  - Dataverse_Import sheet: Client / Site City / Job # label-value rows
"""
import pathlib

import openpyxl
from openpyxl.styles import Font


def build(path: pathlib.Path, job_number: str = "J1-TEST-001",
          scope_name: str = "A1) MV - Test") -> pathlib.Path:
    # job_number sets the Dataverse_Import "Job #" (the parsed project_number); scope_name sets the
    # scope sheet B2 (the line_uid prefix). Both are parameterized so a caller can build a SECOND,
    # distinct-project workbook for multi-project / collision tests. Keep "A_n)" prefix so the scope
    # sheet is detected (SCOPE_RE) and the default reconciliation (2 lines, 25h, P4=1000) holds.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = scope_name[:31]
    ws["B2"] = scope_name
    ws["I3"] = "Total App Hours"
    ws["J3"] = 25.0
    ws["L3"] = "TOTAL SHEET $$$ -NOT ADJUSTED"
    ws["P3"] = 1000.0
    ws["M4"] = 1
    ws["N4"] = 1.0
    ws["L4"] = "TOTAL SHEET $$$ - ADJUSTED"
    ws["P4"] = 1000.0
    ws["C5"] = "QTY"
    ws["D5"] = "Section"
    ws["E5"] = "Apparatus Type"
    ws["G5"] = "Drawing"
    ws["I5"] = "Hrs/Unit"
    ws["J5"] = "Hrs/Line"

    # Bold section header row above apparatus rows (Task 4)
    section_cell = ws["E6"]
    section_cell.value = "Switchgear — Section A"
    section_cell.font = Font(bold=True)

    ws["D7"] = 0
    ws["E7"] = "SLD - sub-header (skip)"  # no QTY -> skipped
    ws["C8"] = 2
    ws["D8"] = "7.6"
    ws["E8"] = "Circuit Breaker MV - Vacuum Bkr"
    ws["G8"] = "E01"
    ws["I8"] = 5.0
    ws["J8"] = 10.0
    ws["C9"] = 3
    ws["D9"] = "7.2"
    ws["E9"] = "Capcitors - Per Unit"
    ws["G9"] = "E01"
    ws["I9"] = 5.0
    ws["J9"] = 15.0
    ws["L14"] = "Onsite Labor Totals"
    ws["P14"] = 800.0
    ws["L19"] = "Offsite Labor Totals"
    ws["P19"] = 100.0
    ws["L26"] = "Travel Totals"
    ws["P26"] = 50.0
    ws["L33"] = "Outside Services Totals"
    ws["P33"] = 50.0

    er = wb.create_sheet("Equipment Reference")  # real layout: B=ATS-sec, C=MTS-sec, D=type, E=ATS25, F=MTS23
    er["B2"] = "ATS"
    er["C2"] = "MTS"
    er["D2"] = "Scope of Work"
    er["E2"] = "ATS25"
    er["F2"] = "MTS23"
    er["B3"] = "7.6"
    er["C3"] = "7.6"
    er["D3"] = "Circuit Breaker MV - Vacuum Bkr"
    er["E3"] = 5.0
    er["F3"] = 5.0

    pt = wb.create_sheet("Print_Template")
    pt["L13"] = "  TOTAL COST"
    pt["R13"] = 1000.0
    pt["L14"] = "  A1) MV - Test"
    pt["R14"] = 1000.0

    # Dataverse_Import metadata sheet (Task 4)
    dv = wb.create_sheet("Dataverse_Import")
    dv["A1"] = "Client:"
    dv["B1"] = "ACME Power Co."
    dv["A2"] = "Site City:"
    dv["B2"] = "Mesa"
    dv["A3"] = "Job #:"
    dv["B3"] = job_number

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


if __name__ == "__main__":
    print(build(pathlib.Path(__file__).parent / "mini_estimator.xlsx"))
