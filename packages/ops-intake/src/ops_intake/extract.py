from __future__ import annotations

import pathlib
import re

import openpyxl

from .model import IntakePayload, ProjectIn, QuoteLineIn, ScopeIn, ScopeQuoteIn, StandardHourIn

SKIP_SHEETS = {"Submittal Specs", "Equipment Reference", "Print_Template", "Dataverse_Import"}
SCOPE_RE = re.compile(r"^[AB]\d\)")


def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):  # the estimator stores some Hrs/Unit / QTY cells as text ("2", "8")
        try:
            return float(v.strip().replace(",", ""))
        except ValueError:
            return None
    return None


def _str(v):
    s = str(v).strip() if v is not None else ""
    return s or None


def _is_bold(cell) -> bool:
    """Return True if the cell has bold formatting (works for ReadOnlyCell in openpyxl 3.x)."""
    try:
        return bool(cell.font and cell.font.b)
    except Exception:
        return False


def _is_scope_sheet(ws) -> bool:
    name = ws["B2"].value
    if not name or ws.title in SKIP_SHEETS or ws.title.endswith(".X"):
        return False
    return bool(SCOPE_RE.match(str(name))) or _num(ws["J3"].value) is not None


def _extract_scope(ws) -> ScopeIn:
    scope_name = str(ws["B2"].value).strip()
    q = ScopeQuoteIn(
        onsite_labor=_num(ws["P14"].value) or 0.0,
        offsite_labor=_num(ws["P19"].value) or 0.0,
        travel=_num(ws["P26"].value) or 0.0,
        outside_services=_num(ws["P33"].value) or 0.0,
        unit_multiplier=_num(ws["M4"].value) or 1.0,
        pct_adjust=_num(ws["N4"].value) if _num(ws["N4"].value) is not None else 1.0,
        total_quoted_hours=_num(ws["J3"].value) or 0.0,
    )
    lines: list[QuoteLineIn] = []
    current_section: str | None = None
    for r in range(6, ws.max_row + 1):
        # Check col E for a bold section-header row (mirrors BuildApparatusJSON macro logic)
        e_cell = ws.cell(r, 5)  # col E
        c_cell = ws.cell(r, 3)  # col C = QTY
        qty = _num(c_cell.value)
        if qty is None or qty <= 0:
            # Might be a section header — check bold on col E
            e_val = _str(e_cell.value)
            if e_val and _is_bold(e_cell):
                current_section = e_val
            continue
        atype = _str(e_cell.value)  # col E = Apparatus Type
        if atype is None:
            continue
        lines.append(QuoteLineIn(
            apparatus_type=atype,
            test_standard="ATS",
            qty=int(round(qty)),
            hrs_per_unit=_num(ws.cell(r, 9).value) or 0.0,  # col I (text or number)
            neta_section=_str(ws.cell(r, 4).value),  # col D
            drawing=_str(ws.cell(r, 7).value),  # col G
            line_number=r,
            section=current_section,
            line_uid=f"{scope_name}:row{r}",
        ))
    return ScopeIn(scope_name=scope_name, quote=q, lines=lines)


def _extract_metadata(wb) -> dict:
    """Read label/value rows from the Dataverse_Import sheet onto a dict for ProjectIn."""
    if "Dataverse_Import" not in wb.sheetnames:
        return {}
    dv = wb["Dataverse_Import"]
    mapping = {}
    for r in range(1, dv.max_row + 1):
        label = _str(dv.cell(r, 1).value)
        value = _str(dv.cell(r, 2).value)
        if not label:
            continue
        label_upper = label.upper().rstrip(":")
        if label_upper == "CLIENT":
            mapping["client_name"] = value
        elif label_upper in ("SITE NAME", "PROJECT"):
            mapping["site_name"] = value
        elif label_upper == "SITE ADDRESS":
            mapping["site_address"] = value
        elif label_upper == "SITE CITY":
            mapping["site_city"] = value
        elif label_upper == "SITE STATE":
            mapping["site_state"] = value
        elif label_upper in ("SITE ZIP", "ZIP"):
            mapping["site_zip"] = value
    return mapping


def _extract_standard_hours(wb) -> list[StandardHourIn]:
    if "Equipment Reference" not in wb.sheetnames:
        return []
    er = wb["Equipment Reference"]
    out: list[StandardHourIn] = []
    for r in range(3, er.max_row + 1):
        sow = _str(er.cell(r, 4).value)  # col D "Scope of Work" (apparatus type)
        ats = _num(er.cell(r, 5).value)  # col E ATS25 hours
        if not sow or ats is None:
            continue
        out.append(StandardHourIn(apparatus_type=sow, test_standard="ATS",
                                  default_hours=ats, neta_section=_str(er.cell(r, 2).value)))  # col B NETA'25
    return out


def _contract_value(wb) -> float:
    if "Print_Template" not in wb.sheetnames:
        return 0.0
    pt = wb["Print_Template"]
    for r in range(1, pt.max_row + 1):
        label = str(pt.cell(r, 12).value or "").strip().upper()  # col L
        if label.startswith("TOTAL COST"):
            return round(_num(pt.cell(r, 18).value) or 0.0, 2)  # col R
    return 0.0


def _extract_chiller_scopes(wb, start_sort: int) -> list[ScopeIn]:
    """Chiller scopes appear only in the Print_Template rollup (no scope sheet); capture the lump
    as a flagged estimate (single category) so the project total reconciles."""
    if "Print_Template" not in wb.sheetnames:
        return []
    pt = wb["Print_Template"]
    out: list[ScopeIn] = []
    for r in range(1, pt.max_row + 1):
        label = _str(pt.cell(r, 12).value)  # col L
        if not label or "chiller" not in label.lower():
            continue
        amt = _num(pt.cell(r, 18).value) or 0.0  # col R
        out.append(ScopeIn(
            scope_name=label.lstrip(" *").strip(),
            sort_order=start_sort + len(out),
            quote=ScopeQuoteIn(outside_services=amt, is_estimate=True),
            lines=[],
        ))
    return out


def extract_workbook(path) -> IntakePayload:
    path = pathlib.Path(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    scopes: list[ScopeIn] = []
    for ws in wb.worksheets:
        if _is_scope_sheet(ws):
            s = _extract_scope(ws)
            s.sort_order = len(scopes) + 1
            scopes.append(s)
    scopes.extend(_extract_chiller_scopes(wb, len(scopes) + 1))
    metadata = _extract_metadata(wb)
    project = ProjectIn(
        project_number="MINER-PHX-AB-MV",
        project_name="Project Miner — PHX Bldg A & B MV",
        status="Won",
        quote_revision="Rev10",
        contract_value=_contract_value(wb),
        description=("Public/product name: Project Jupiter — Oracle/STACK data-center campus, "
                     "Doña Ana County NM."),
        client_name=metadata.get("client_name"),
        site_name=metadata.get("site_name"),
        site_address=metadata.get("site_address"),
        site_city=metadata.get("site_city"),
        site_state=metadata.get("site_state"),
        site_zip=metadata.get("site_zip"),
    )
    return IntakePayload(project=project, scopes=scopes, standard_hours=_extract_standard_hours(wb))
