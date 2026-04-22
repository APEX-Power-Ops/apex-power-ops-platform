"""
MOP Work Script — Excel (.xlsx) Template Generator
====================================================
Generates an editable Excel workbook matching the HTML work-script table design.
Includes RESA brand styling, phase headers, risk badges, and placeholder fields.

Usage:
    python Scripts/generate_work_script_xlsx.py

Output:
    Templates/MOP/MOP_Work_Script_Template.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter

# ── Brand Colors ──
RESA_BLUE = "015687"
RESA_GREEN = "5FA844"
RESA_GRAY = "686866"
LIGHT_GRAY = "F5F5F5"
BORDER_GRAY = "DDDDDD"
PHASE_BG = "C9A227"
PHASE_TEXT = "1A1A1A"
RISK_HIGH = "DC3545"
RISK_MED = "FFC107"
RISK_LOW = "28A745"
WHITE = "FFFFFF"
BLACK = "333333"
AMBER_LIGHT = "FFF8E6"

# ── Paths ──
PACKAGE_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_PATH = os.path.join(PACKAGE_ROOT, "artifacts", "MOP_Work_Script_Template.xlsx")

# ── Work Script Data (from HTML) ──
PHASES = [
    {
        "name": "Pre-Work",
        "steps": [
            {
                "step": 1,
                "procedure": "Conduct safety briefing with all personnel. Review MOP, hazards, PPE requirements, and emergency procedures.",
                "resources": "RESA Power\nCustomer Rep",
                "est": "15m",
                "risk": "Low",
                "hold": True,
                "note": "",
                "cease": "",
            },
            {
                "step": 2,
                "procedure": "Verify all approvals, permits, and clearances in place.",
                "resources": "RESA Power\nCustomer Rep",
                "est": "10m",
                "risk": "Low",
                "hold": True,
                "note": "",
                "cease": "",
            },
            {
                "step": 3,
                "procedure": "Contact Control Room. Confirm readiness to proceed.",
                "resources": "RESA Power\nControl Room",
                "est": "10m",
                "risk": "Low",
                "hold": True,
                "note": "Document clearance # and operator name",
                "cease": "",
            },
        ],
    },
    {
        "name": "Isolation",
        "steps": [
            {
                "step": 4,
                "procedure": "Perform switching/isolation per approved switching order.",
                "resources": "Electrical Contractor\nRESA Power",
                "est": "30m",
                "risk": "High",
                "hold": False,
                "note": "Ref: SOP-ISO-001",
                "cease": "Cease testing if anomaly occurs",
            },
            {
                "step": 5,
                "procedure": "Apply LOTO devices at all isolation points. Verify all personnel locks applied.",
                "resources": "RESA Power\nAll Personnel",
                "est": "15m",
                "risk": "Med",
                "hold": False,
                "note": "",
                "cease": "",
            },
            {
                "step": 6,
                "procedure": "Perform voltage absence verification (3-point method).",
                "resources": "RESA Power\nCustomer Witness",
                "est": "15m",
                "risk": "High",
                "hold": True,
                "note": "Form: IVF-001",
                "cease": "Cease testing if anomaly occurs",
            },
            {
                "step": 7,
                "procedure": "Install Temporary Protective Grounds per SOP-GRD-001.",
                "resources": "RESA Power",
                "est": "20m",
                "risk": "Med",
                "hold": False,
                "note": "Form: IVF-001",
                "cease": "",
            },
        ],
    },
    {
        "name": "Execution",
        "steps": [
            {
                "step": 8,
                "procedure": "Remove existing protective relays. Document wiring/settings.",
                "resources": "RESA Power",
                "est": "1h",
                "risk": "Med",
                "hold": False,
                "note": "Ref: SOP-REM-001",
                "cease": "",
            },
            {
                "step": 9,
                "procedure": "Install new protective relays.",
                "resources": "RESA Power\nVendor (if req'd)",
                "est": "1.5h",
                "risk": "Med",
                "hold": False,
                "note": "Ref: SOP-INST-001",
                "cease": "",
            },
            {
                "step": 10,
                "procedure": "Configure relay settings per protection study.",
                "resources": "RESA Power",
                "est": "1h",
                "risk": "Med",
                "hold": False,
                "note": "Ref: SOP-TEST-001",
                "cease": "",
            },
            {
                "step": 11,
                "procedure": "Perform functional and protection testing per NETA ATS.",
                "resources": "RESA Power\nCustomer Witness",
                "est": "2h",
                "risk": "Med",
                "hold": False,
                "note": "Form: FTR-001",
                "cease": "",
            },
        ],
    },
    {
        "name": "Verification",
        "steps": [
            {
                "step": 12,
                "procedure": "Verify all work complete. No tools/debris in equipment.",
                "resources": "RESA Power\nCustomer Rep",
                "est": "15m",
                "risk": "Low",
                "hold": True,
                "note": "",
                "cease": "",
            },
        ],
    },
    {
        "name": "Restoration",
        "steps": [
            {
                "step": 13,
                "procedure": "Remove TPGs (line end first, ground end last).",
                "resources": "RESA Power",
                "est": "15m",
                "risk": "High",
                "hold": False,
                "note": "",
                "cease": "Cease testing if anomaly occurs",
            },
            {
                "step": 14,
                "procedure": "Notify Control Room - ready to restore. Obtain LOTO removal auth.",
                "resources": "RESA Power\nControl Room",
                "est": "10m",
                "risk": "Med",
                "hold": False,
                "note": "",
                "cease": "",
            },
            {
                "step": 15,
                "procedure": "Remove all LOTO devices. Account for all locks.",
                "resources": "RESA Power\nAll Personnel",
                "est": "10m",
                "risk": "Med",
                "hold": False,
                "note": "",
                "cease": "",
            },
            {
                "step": 16,
                "procedure": "Perform restoration switching per approved order.",
                "resources": "Electrical Contractor\nRESA Power\nControl Room",
                "est": "30m",
                "risk": "High",
                "hold": False,
                "note": "Ref: SOP-ISO-001",
                "cease": "Cease testing if anomaly occurs",
            },
            {
                "step": 17,
                "procedure": "Verify equipment energized and operating normally.",
                "resources": "RESA Power\nCustomer Rep",
                "est": "15m",
                "risk": "Med",
                "hold": False,
                "note": "",
                "cease": "",
            },
        ],
    },
    {
        "name": "Closeout",
        "steps": [
            {
                "step": 18,
                "procedure": "Notify Control Room - work complete. Release clearance.",
                "resources": "RESA Power\nControl Room",
                "est": "10m",
                "risk": "Low",
                "hold": False,
                "note": "",
                "cease": "",
            },
            {
                "step": 19,
                "procedure": "Complete all documentation. Obtain customer sign-off.",
                "resources": "RESA Power\nCustomer Rep",
                "est": "15m",
                "risk": "Low",
                "hold": True,
                "note": "",
                "cease": "",
            },
        ],
    },
]


# ── Styles ──
THIN_BORDER = Side(style="thin", color=BORDER_GRAY)
CELL_BORDER = Border(
    left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER
)

HEADER_FONT = Font(name="Segoe UI", size=8, bold=True, color=WHITE)
HEADER_FILL = PatternFill(start_color=RESA_BLUE, end_color=RESA_BLUE, fill_type="solid")
HEADER_BORDER = Border(
    left=Side(style="thin", color=RESA_BLUE),
    right=Side(style="thin", color=RESA_BLUE),
    top=Side(style="thin", color=RESA_BLUE),
    bottom=Side(style="thin", color=RESA_BLUE),
)

PHASE_FONT = Font(name="Segoe UI", size=9, bold=True, color=PHASE_TEXT)
PHASE_FILL = PatternFill(start_color=PHASE_BG, end_color=PHASE_BG, fill_type="solid")

BODY_FONT = Font(name="Segoe UI", size=9, color=BLACK)
NOTE_FONT = Font(name="Segoe UI", size=8, italic=True, color=RESA_BLUE)
CEASE_FONT = Font(name="Segoe UI", size=8, bold=True, color=RISK_HIGH)
HOLD_FONT = Font(name="Segoe UI", size=8, bold=True, color=PHASE_TEXT)
RESOURCE_FONT = Font(name="Segoe UI", size=8, color=RESA_GRAY)

RISK_FILLS = {
    "Low": PatternFill(start_color=RISK_LOW, end_color=RISK_LOW, fill_type="solid"),
    "Med": PatternFill(start_color=RISK_MED, end_color=RISK_MED, fill_type="solid"),
    "High": PatternFill(start_color=RISK_HIGH, end_color=RISK_HIGH, fill_type="solid"),
}
RISK_FONTS = {
    "Low": Font(name="Segoe UI", size=8, bold=True, color=WHITE),
    "Med": Font(name="Segoe UI", size=8, bold=True, color=BLACK),
    "High": Font(name="Segoe UI", size=8, bold=True, color=WHITE),
}


def build_workbook():
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Work Script"

    # ── Column Widths (approximate HTML percentages for letter page) ──
    col_widths = {
        "A": 5,     # # (step)
        "B": 48,    # Procedure
        "C": 16,    # Resources
        "D": 7,     # Est.
        "E": 10,    # Start
        "F": 10,    # End
        "G": 7,     # Risk
        "H": 7,     # Init
        "I": 7,     # Ver
        "J": 20,    # Notes (bonus not in HTML - captures proc-notes)
        "K": 22,    # Warnings (bonus - captures cease-warnings)
        "L": 7,     # Hold Point
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── Title Row ──
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    title_cell = ws.cell(row=row, column=1, value="METHOD OF PROCEDURE — WORK SCRIPT")
    title_cell.font = Font(name="Segoe UI", size=14, bold=True, color=RESA_BLUE)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 30

    # ── Document Info Row ──
    row = 2
    info_fields = [
        ("MOP ID:", "{{mop_id}}"),
        ("Version:", "{{version}}"),
        ("Project:", "{{project_name}}"),
        ("Site:", "{{site_name}}"),
        ("Date:", "{{date}}"),
    ]
    col = 1
    for label, value in info_fields:
        cell_l = ws.cell(row=row, column=col, value=label)
        cell_l.font = Font(name="Segoe UI", size=8, bold=True, color=RESA_GRAY)
        cell_l.alignment = Alignment(horizontal="right")
        col += 1
        cell_v = ws.cell(row=row, column=col, value=value)
        cell_v.font = Font(name="Segoe UI", size=9, color=BLACK)
        col += 1
    ws.row_dimensions[row].height = 20

    # ── Spacer ──
    row = 3
    ws.row_dimensions[row].height = 6

    # ── Column Headers ──
    row = 4
    headers = ["#", "Procedure", "Resources", "Est.", "Start", "End", "Risk", "Init", "Ver", "Notes", "Warnings", "Hold"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(
            horizontal="center" if col_idx != 2 else "left",
            vertical="center",
            wrap_text=True,
        )
    ws.row_dimensions[row].height = 22

    # ── Step Data ──
    row = 5
    for phase in PHASES:
        # Phase header row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        phase_cell = ws.cell(row=row, column=1, value=phase["name"].upper())
        phase_cell.font = PHASE_FONT
        phase_cell.fill = PHASE_FILL
        phase_cell.alignment = Alignment(horizontal="left", vertical="center")
        phase_cell.border = Border(
            left=Side(style="thin", color=PHASE_BG),
            right=Side(style="thin", color=PHASE_BG),
            top=Side(style="thin", color=PHASE_BG),
            bottom=Side(style="thin", color=PHASE_BG),
        )
        ws.row_dimensions[row].height = 20
        row += 1

        for step_data in phase["steps"]:
            # Build procedure text (with hold indicator inline)
            proc_text = step_data["procedure"]
            if step_data["hold"]:
                proc_text += "  [HOLD POINT]"

            # Step #
            cell = ws.cell(row=row, column=1, value=step_data["step"])
            cell.font = Font(name="Segoe UI", size=9, bold=True, color=BLACK)
            cell.alignment = Alignment(horizontal="center", vertical="top")
            cell.border = CELL_BORDER

            # Procedure
            cell = ws.cell(row=row, column=2, value=proc_text)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = CELL_BORDER

            # Resources
            cell = ws.cell(row=row, column=3, value=step_data["resources"])
            cell.font = RESOURCE_FONT
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = CELL_BORDER

            # Est.
            cell = ws.cell(row=row, column=4, value=step_data["est"])
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center", vertical="top")
            cell.border = CELL_BORDER

            # Start (blank for input)
            cell = ws.cell(row=row, column=5, value="")
            cell.border = CELL_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="top")

            # End (blank for input)
            cell = ws.cell(row=row, column=6, value="")
            cell.border = CELL_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="top")

            # Risk (color-coded)
            risk = step_data["risk"]
            cell = ws.cell(row=row, column=7, value=risk)
            cell.font = RISK_FONTS.get(risk, BODY_FONT)
            cell.fill = RISK_FILLS.get(risk, PatternFill())
            cell.alignment = Alignment(horizontal="center", vertical="top")
            cell.border = CELL_BORDER

            # Init (blank for input)
            cell = ws.cell(row=row, column=8, value="")
            cell.border = CELL_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="top")

            # Ver (blank for input)
            cell = ws.cell(row=row, column=9, value="")
            cell.border = CELL_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="top")

            # Notes
            cell = ws.cell(row=row, column=10, value=step_data.get("note", ""))
            cell.font = NOTE_FONT
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = CELL_BORDER

            # Warnings
            cell = ws.cell(row=row, column=11, value=step_data.get("cease", ""))
            if step_data.get("cease"):
                cell.font = CEASE_FONT
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = CELL_BORDER

            # Hold
            hold_val = "YES" if step_data["hold"] else ""
            cell = ws.cell(row=row, column=12, value=hold_val)
            if step_data["hold"]:
                cell.font = HOLD_FONT
                cell.fill = PatternFill(start_color=AMBER_LIGHT, end_color=AMBER_LIGHT, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="top")
            cell.border = CELL_BORDER

            # Row height based on procedure text length
            proc_len = len(proc_text)
            ws.row_dimensions[row].height = max(28, min(55, proc_len // 2))

            row += 1

    # ── Summary Footer ──
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    footer_cell = ws.cell(row=row, column=1, value="RESA Power Service  ·  NETA Accredited Company  ·  resapower.com")
    footer_cell.font = Font(name="Segoe UI", size=8, color=RESA_GRAY)
    footer_cell.alignment = Alignment(horizontal="center")

    # ── Print Setup ──
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # Freeze panes (header row)
    ws.freeze_panes = "A5"

    # ── Save ──
    wb.save(OUTPUT_PATH)
    print(f"✅ Excel template saved: {OUTPUT_PATH}")
    print(f"   File size: {os.path.getsize(OUTPUT_PATH) / 1024:.1f} KB")


if __name__ == "__main__":
    build_workbook()
